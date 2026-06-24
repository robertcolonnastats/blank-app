"""
mets_milb_lib.py

Shared logic for pulling Mets minor-league stats (box score, all levels) +
Statcast (Triple-A & Single-A only, where it's publicly tracked) from the
MLB Stats API, and writing a sortable Excel workbook.

Used by both:
  - pull_and_save.py (run on a schedule via GitHub Actions, no UI)
  - streamlit_app.py (manual on-demand pull + viewing the latest saved file)
"""

import io
import time

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = "https://statsapi.mlb.com/api/v1"
METS_ORG_ID = 121

SPORT_IDS = {
    11: "Triple-A",
    12: "Double-A",
    13: "High-A",
    14: "Single-A",
    16: "Rookie/Complex (FCL)",
    5442: "Dominican Summer League",
}

STATCAST_TRACKED_LEVELS = {"Triple-A", "Single-A"}

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "personal-research-app/1.0"})


# ---------------------------------------------------------------------------
# Low-level API helpers
# ---------------------------------------------------------------------------

def get_json(url, params=None, retries=3, sleep=0.4):
    for attempt in range(retries):
        try:
            r = SESSION.get(url, params=params, timeout=20)
            r.raise_for_status()
            return r.json()
        except Exception:
            if attempt == retries - 1:
                return {}
            time.sleep(sleep)
    return {}


def get_affiliate_teams(season):
    teams = []
    for sport_id, level_name in SPORT_IDS.items():
        data = get_json(f"{BASE}/teams", params={"sportId": sport_id, "season": season})
        for t in data.get("teams", []):
            if t.get("parentOrgId") == METS_ORG_ID:
                teams.append({
                    "teamId": t["id"],
                    "teamName": t.get("name"),
                    "sportId": sport_id,
                    "levelName": level_name,
                })
    return teams


def get_team_group_stats(team_id, sport_id, season, group):
    data = get_json(f"{BASE}/stats", params={
        "stats": "season",
        "group": group,
        "season": season,
        "sportId": sport_id,
        "teamId": team_id,
        "limit": 300,
    })
    rows = []
    for block in data.get("stats", []):
        for split in block.get("splits", []):
            player = split.get("player", {})
            stat = split.get("stat", {})
            row = {"playerId": player.get("id"), "playerName": player.get("fullName")}
            row.update(stat)
            rows.append(row)
    return rows


def get_team_fielding_stats(team_id, sport_id, season):
    """Box-score-derived fielding (putouts, assists, errors, fielding%,
    innings, position). This is the only public defensive data that
    exists for minor leaguers -- no OAA/DRS equivalent is published below
    MLB, so any Field grade built from this is necessarily a crude proxy,
    not a true range/glove evaluation. A player can have multiple rows
    here if they played multiple positions; caller should aggregate."""
    rows = get_team_group_stats(team_id, sport_id, season, "fielding")
    return rows


def get_player_bio(player_id):
    data = get_json(f"{BASE}/people/{player_id}")
    people = data.get("people", [])
    if not people:
        return {}
    p = people[0]
    return {
        "birthDate": p.get("birthDate"),
        "currentAge": p.get("currentAge"),
        "position": (p.get("primaryPosition") or {}).get("abbreviation"),
        "bats": (p.get("batSide") or {}).get("code"),
        "throws": (p.get("pitchHand") or {}).get("code"),
        "height": p.get("height"),
        "weight": p.get("weight"),
    }


def get_team_game_pks(team_id, sport_id, season):
    data = get_json(f"{BASE}/schedule", params={
        "teamId": team_id,
        "sportId": sport_id,
        "season": season,
        "gameType": "R",
    })
    pks = []
    for date_block in data.get("dates", []):
        for game in date_block.get("games", []):
            if game.get("status", {}).get("abstractGameState") == "Final":
                pks.append(game["gamePk"])
    return pks


def get_game_events(game_pk):
    """Single pass through a game's play-by-play, returning two lists:
    batted_balls (hitData on contact events) and pitches (every tracked
    pitch: type, velocity, spin, movement, call result, zone). Pulling
    both from one fetch avoids hitting the API twice per game."""
    data = get_json(f"{BASE}/game/{game_pk}/playByPlay")
    batted_balls, pitches = [], []
    for play in data.get("allPlays", []):
        matchup = play.get("matchup", {})
        batter = matchup.get("batter", {})
        pitcher = matchup.get("pitcher", {})
        for event in play.get("playEvents", []):
            if not event.get("isPitch"):
                continue
            details = event.get("details", {})
            pitch_data = event.get("pitchData", {}) or {}
            breaks = pitch_data.get("breaks", {}) or {}
            coords = pitch_data.get("coordinates", {}) or {}
            pitch_type_info = details.get("type") or {}

            pitches.append({
                "pitcherId": pitcher.get("id"),
                "pitcherName": pitcher.get("fullName"),
                "batterId": batter.get("id"),
                "batterName": batter.get("fullName"),
                "pitch_type": pitch_type_info.get("code"),
                "pitch_desc": pitch_type_info.get("description"),
                "start_speed": pitch_data.get("startSpeed"),
                "spin_rate": breaks.get("spinRate"),
                "pfx_x": coords.get("pfxX"),
                "pfx_z": coords.get("pfxZ"),
                "call_description": details.get("description"),
                "zone": pitch_data.get("zone"),
            })

            hit_data = event.get("hitData")
            if hit_data and hit_data.get("launchSpeed") is not None:
                batted_balls.append({
                    "playerId": batter.get("id"),
                    "playerName": batter.get("fullName"),
                    "pitcherId": pitcher.get("id"),
                    "pitcherName": pitcher.get("fullName"),
                    "launch_speed": hit_data.get("launchSpeed"),
                    "launch_angle": hit_data.get("launchAngle"),
                    "total_distance": hit_data.get("totalDistance"),
                })
    return batted_balls, pitches


def pull_game_data_for_team(team_id, sport_id, season, progress_callback=None, label=""):
    """Pulls every game for a team and returns (batted_balls_df,
    pitches_df). Runs for ALL levels -- whether a level actually has
    tracking data is something the response itself will reveal (empty
    columns / NaNs), rather than something we assume up front."""
    game_pks = get_team_game_pks(team_id, sport_id, season)
    all_bb, all_pitches = [], []
    for i, pk in enumerate(game_pks):
        if progress_callback and i % 5 == 0:
            progress_callback(f"{label}: game {i + 1}/{len(game_pks)}...")
        bb, pitches = get_game_events(pk)
        all_bb.extend(bb)
        all_pitches.extend(pitches)
        time.sleep(0.1)
    return pd.DataFrame(all_bb), pd.DataFrame(all_pitches)


# ---------------------------------------------------------------------------
# Derived stats
# ---------------------------------------------------------------------------

def safe_div(n, d):
    try:
        n = float(n)
        d = float(d)
        return n / d if d else None
    except (TypeError, ValueError):
        return None


def add_hitting_derived(df):
    if df.empty:
        return df
    for col in ["hits", "homeRuns", "atBats", "strikeOuts", "baseOnBalls",
                "sacFlies", "plateAppearances"]:
        if col not in df.columns:
            df[col] = 0
    df["BABIP"] = df.apply(
        lambda r: safe_div(
            (float(r["hits"] or 0) - float(r["homeRuns"] or 0)),
            (float(r["atBats"] or 0) - float(r["strikeOuts"] or 0)
             - float(r["homeRuns"] or 0) + float(r["sacFlies"] or 0)),
        ), axis=1)
    df["K_pct"] = df.apply(lambda r: safe_div(r["strikeOuts"], r["plateAppearances"]), axis=1)
    df["BB_pct"] = df.apply(lambda r: safe_div(r["baseOnBalls"], r["plateAppearances"]), axis=1)
    if "slg" in df.columns and "avg" in df.columns:
        df["ISO"] = df.apply(
            lambda r: (float(r["slg"]) - float(r["avg"]))
            if r.get("slg") not in (None, "", "-") and r.get("avg") not in (None, "", "-")
            else None, axis=1)
    return df


def add_pitching_derived(df):
    if df.empty:
        return df
    for col in ["strikeOuts", "baseOnBalls", "battersFaced"]:
        if col not in df.columns:
            df[col] = 0
    df["K_pct"] = df.apply(lambda r: safe_div(r["strikeOuts"], r["battersFaced"]), axis=1)
    df["BB_pct"] = df.apply(lambda r: safe_div(r["baseOnBalls"], r["battersFaced"]), axis=1)
    df["K_minus_BB_pct"] = df.apply(
        lambda r: (r["K_pct"] - r["BB_pct"]) if r["K_pct"] is not None and r["BB_pct"] is not None else None,
        axis=1)
    return df


# ---------------------------------------------------------------------------
# 20-80 scouting grades (Hit / Power / Field / Run / Overall FV)
#
# Honest data-coverage notes, since this is presented as a "grade" and
# shouldn't imply more certainty than the underlying data supports:
#   - Hit, Power: built from real statistical signal (contact rate, AVG/
#     ISO, and Statcast contact quality where tracked). Reasonable
#     confidence.
#   - Field: built from box-score fielding stats only (fielding%, range
#     factor). No OAA/DRS-equivalent exists publicly for the minors, so
#     this is a crude proxy, not a true defensive evaluation. Low-medium
#     confidence.
#   - Run: built ONLY from stolen-base attempt/success rate, since there
#     is no public sprint-speed data for minor leaguers anywhere (MLB
#     tracks it internally via Hawk-Eye but does not publish it below the
#     majors). A fast player who never attempts steals looks identical to
#     a slow one. Low confidence -- shown but clearly flagged.
#   - Arm: NO public data exists for minor leaguers at any level. Left as
#     None/blank everywhere rather than guessed.
# ---------------------------------------------------------------------------

GRADE_WEIGHTS = {"hit": 0.30, "power": 0.25, "field": 0.20, "run": 0.15, "arm": 0.10}


def to_20_80(z):
    if z is None or pd.isna(z):
        return None
    val = 50 + (z * 10)
    val = max(20, min(80, val))
    return int(round(val / 5.0) * 5)


def aggregate_fielding(raw_rows):
    """Collapse multi-position fielding rows into one row per player:
    primary position (most innings) + summed counting stats."""
    if not raw_rows:
        return pd.DataFrame()
    df = pd.DataFrame(raw_rows)
    for col in ["putOuts", "assists", "errors", "chances", "innings"]:
        if col not in df.columns:
            df[col] = 0
    df["innings_num"] = pd.to_numeric(df["innings"], errors="coerce").fillna(0)

    def agg(g):
        primary = g.loc[g["innings_num"].idxmax()] if not g.empty else g.iloc[0]
        po = pd.to_numeric(g["putOuts"], errors="coerce").fillna(0).sum()
        ast = pd.to_numeric(g["assists"], errors="coerce").fillna(0).sum()
        err = pd.to_numeric(g["errors"], errors="coerce").fillna(0).sum()
        innings_total = g["innings_num"].sum()
        chances = po + ast + err
        fielding_pct = (po + ast) / chances if chances else None
        range_factor_9 = ((po + ast) / innings_total * 9) if innings_total else None
        return pd.Series({
            "primary_position": primary.get("position", {}).get("abbreviation")
            if isinstance(primary.get("position"), dict) else primary.get("position"),
            "fielding_pct": fielding_pct,
            "range_factor_9": range_factor_9,
            "innings_fielded": innings_total,
        })

    return df.groupby("playerId").apply(agg).reset_index()


def compute_tool_grades(hitting):
    """Adds hit_grade/power_grade/field_grade/run_grade/arm_grade
    (20-80) and overall_fv to the hitting dataframe. Arm is always None.
    Overall FV only uses whichever tools are actually present, with
    weights renormalized over the available subset (Arm is always
    excluded since it's never available)."""
    if hitting.empty:
        return hitting

    df = hitting.copy()

    # --- Hit: contact rate + AVG, plus real swing-decision data (chase%,
    # whiff%, zone-contact%) where pitch-level zone/call data is tracked,
    # plus Statcast contact quality where that's tracked too.
    df["_z_contact"] = -df.groupby("level")["K_pct"].transform(zscore)
    df["_z_avg"] = df.groupby("level")["avg"].transform(lambda s: zscore(pd.to_numeric(s, errors="coerce")))
    hit_cols = ["_z_contact", "_z_avg"]
    has_approach_data = "chase_pct" in df.columns
    if has_approach_data:
        df["_z_chase_inv"] = -df.groupby("level")["chase_pct"].transform(zscore)
        df["_z_bwhiff_inv"] = -df.groupby("level")["batter_whiff_pct"].transform(zscore)
        df["_z_zonecontact"] = df.groupby("level")["zone_contact_pct"].transform(zscore)
        hit_cols += ["_z_chase_inv", "_z_bwhiff_inv", "_z_zonecontact"]
    if "avg_exit_velocity" in df.columns:
        df["_z_ev_hit"] = df.groupby("level")["avg_exit_velocity"].transform(zscore)
        df["_z_hardhit_hit"] = df.groupby("level")["hard_hit_pct"].transform(zscore)
        hit_cols += ["_z_ev_hit", "_z_hardhit_hit"]
    df["_hit_z"] = df.apply(lambda r: _row_mean_ignore_none(r, hit_cols), axis=1)
    df["hit_grade"] = df["_hit_z"].apply(to_20_80)
    df["hit_grade_confidence"] = (
        "medium-high (contact rate, AVG, chase%, whiff%, zone-contact%"
        + (", Statcast contact quality)" if "avg_exit_velocity" in df.columns else ")")
        if has_approach_data else
        "medium (contact rate + AVG only -- no pitch-level swing-decision data at this level; "
        "still no quality-of-competition adjustment, since that data doesn't exist publicly anywhere)"
    )

    # --- Power: ISO + HR rate, plus Statcast max EV/barrel% if tracked
    df["_hr_rate"] = df.apply(lambda r: safe_div(r.get("homeRuns"), r.get("plateAppearances")), axis=1)
    df["_z_iso"] = df.groupby("level")["ISO"].transform(lambda s: zscore(pd.to_numeric(s, errors="coerce")))
    df["_z_hrrate"] = df.groupby("level")["_hr_rate"].transform(zscore)
    power_cols = ["_z_iso", "_z_hrrate"]
    if "max_exit_velocity" in df.columns:
        df["_z_maxev"] = df.groupby("level")["max_exit_velocity"].transform(zscore)
        df["_z_barrel_power"] = df.groupby("level")["barrel_pct_approx"].transform(zscore)
        power_cols += ["_z_maxev", "_z_barrel_power"]
    df["_power_z"] = df.apply(lambda r: _row_mean_ignore_none(r, power_cols), axis=1)
    df["power_grade"] = df["_power_z"].apply(to_20_80)

    # --- Field: box-score fielding only, position+level peer group
    if "fielding_pct" in df.columns:
        group_cols = ["level", "primary_position"]
        df["_z_fpct"] = df.groupby(group_cols)["fielding_pct"].transform(
            lambda s: zscore(pd.to_numeric(s, errors="coerce")))
        df["_z_rf"] = df.groupby(group_cols)["range_factor_9"].transform(
            lambda s: zscore(pd.to_numeric(s, errors="coerce")))
        df["_field_z"] = df.apply(lambda r: _row_mean_ignore_none(r, ["_z_fpct", "_z_rf"]), axis=1)
        df["field_grade"] = df["_field_z"].apply(to_20_80)
        df["field_grade_confidence"] = "low-medium (box-score fielding only, no OAA-equivalent exists for minors)"
    else:
        df["field_grade"] = None
        df["field_grade_confidence"] = "unavailable (no fielding data merged)"

    # --- Run: SB rate proxy ONLY -- explicitly low confidence
    if "stolenBases" in df.columns:
        df["_sb_opp"] = df.apply(
            lambda r: safe_div(
                (float(r.get("stolenBases") or 0) + float(r.get("caughtStealing") or 0)),
                (float(r.get("hits") or 0) + float(r.get("baseOnBalls") or 0) + float(r.get("hitByPitch") or 0)),
            ), axis=1)
        df["_sb_success"] = df.apply(
            lambda r: safe_div(r.get("stolenBases"),
                                (float(r.get("stolenBases") or 0) + float(r.get("caughtStealing") or 0)))
            if (float(r.get("stolenBases") or 0) + float(r.get("caughtStealing") or 0)) >= 3 else None,
            axis=1)
        df["_z_sbopp"] = df.groupby("level")["_sb_opp"].transform(zscore)
        df["_z_sbsuccess"] = df.groupby("level")["_sb_success"].transform(zscore)
        df["_run_z"] = df.apply(lambda r: _row_mean_ignore_none(r, ["_z_sbopp", "_z_sbsuccess"]), axis=1)
        df["run_grade"] = df["_run_z"].apply(to_20_80)
        df["run_grade_confidence"] = "low (stolen-base rate only -- no public sprint speed exists for minors)"
    else:
        df["run_grade"] = None
        df["run_grade_confidence"] = "unavailable"

    # --- Arm: no public data exists for minor leaguers anywhere. Always
    # blank rather than guessed.
    df["arm_grade"] = None
    df["arm_grade_confidence"] = "no public data exists for minor leaguers (Arm Strength leaderboard is MLB-only)"

    # --- Overall FV: weighted average of whichever tools are present,
    # weights renormalized over the available subset. Arm is always
    # excluded from the renormalization since it's never available.
    def overall_fv(row):
        available = {
            tool: row.get(f"{tool}_grade")
            for tool in ["hit", "power", "field", "run"]
            if pd.notna(row.get(f"{tool}_grade"))
        }
        if not available or "hit" not in available or "power" not in available:
            return None  # don't produce an FV without at least the two core tools
        total_weight = sum(GRADE_WEIGHTS[t] for t in available)
        return round(sum(GRADE_WEIGHTS[t] * v for t, v in available.items()) / total_weight)

    df["overall_fv"] = df.apply(overall_fv, axis=1)

    df = df.drop(columns=[c for c in df.columns if c.startswith("_")])
    return df


def zscore(series):
    s = pd.to_numeric(series, errors="coerce")
    if s.dropna().empty or s.std(skipna=True) in (0, None) or pd.isna(s.std(skipna=True)):
        return pd.Series([float("nan")] * len(series), index=series.index, dtype="float64")
    return (s - s.mean(skipna=True)) / s.std(skipna=True)


def _row_mean_ignore_none(row, cols):
    import numpy as np
    vals = [row[c] for c in cols if c in row and pd.notna(row[c])]
    return float(np.mean(vals)) if vals else None


def compute_hitting_scores(df):
    """Skill signal: plate discipline + age-for-level + (if available)
    contact quality from Statcast (hard-hit%, barrel%, avg EV).
    Results signal: what the stat line actually shows (AVG/OBP/SLG).
    undervalued_score = skill_z - results_z -- positive means the
    underlying indicators are better than the surface results suggest."""
    if df.empty:
        return df

    df = df.copy()
    # Skill components, z-scored within level so a young/disciplined
    # Double-A guy is compared to Double-A peers, not Triple-A ones.
    df["_z_BB_pct"] = df.groupby("level")["BB_pct"].transform(zscore)
    df["_z_K_pct_inv"] = -df.groupby("level")["K_pct"].transform(zscore)
    df["_z_age_inv"] = -df.groupby("level")["currentAge"].transform(zscore)

    skill_cols = ["_z_BB_pct", "_z_K_pct_inv", "_z_age_inv"]

    if "avg_exit_velocity" in df.columns:
        df["_z_ev"] = df.groupby("level")["avg_exit_velocity"].transform(zscore)
        df["_z_hard_hit"] = df.groupby("level")["hard_hit_pct"].transform(zscore)
        df["_z_barrel"] = df.groupby("level")["barrel_pct_approx"].transform(zscore)
        skill_cols += ["_z_ev", "_z_hard_hit", "_z_barrel"]

    # Results: the actual triple-slash the stat line shows.
    df["_z_avg"] = df.groupby("level")["avg"].transform(lambda s: zscore(pd.to_numeric(s, errors="coerce")))
    df["_z_obp"] = df.groupby("level")["obp"].transform(lambda s: zscore(pd.to_numeric(s, errors="coerce")))
    df["_z_slg"] = df.groupby("level")["slg"].transform(lambda s: zscore(pd.to_numeric(s, errors="coerce")))
    results_cols = ["_z_avg", "_z_obp", "_z_slg"]

    df["skill_z"] = df.apply(lambda r: _row_mean_ignore_none(r, skill_cols), axis=1)
    df["results_z"] = df.apply(lambda r: _row_mean_ignore_none(r, results_cols), axis=1)
    df["undervalued_score"] = df.apply(
        lambda r: (r["skill_z"] - r["results_z"])
        if r["skill_z"] is not None and r["results_z"] is not None else None,
        axis=1)
    df["has_statcast"] = "avg_exit_velocity" in df.columns and df["avg_exit_velocity"].notna()

    df = df.drop(columns=[c for c in df.columns if c.startswith("_z_")])
    return df


def compute_pitching_scores(df):
    """Skill signal: K-BB% + age-for-level + (if available) contact
    quality ALLOWED from Statcast -- lower EV/hard-hit/barrel allowed is
    better, so those are inverted before averaging in.
    Results signal: ERA (inverted, since lower ERA = better results).
    undervalued_score = skill_z - results_z."""
    if df.empty:
        return df

    df = df.copy()
    df["_z_kbb"] = df.groupby("level")["K_minus_BB_pct"].transform(zscore)
    df["_z_age_inv"] = -df.groupby("level")["currentAge"].transform(zscore)
    skill_cols = ["_z_kbb", "_z_age_inv"]

    if "avg_exit_velocity_allowed" in df.columns:
        df["_z_ev_allowed_inv"] = -df.groupby("level")["avg_exit_velocity_allowed"].transform(zscore)
        df["_z_hard_hit_allowed_inv"] = -df.groupby("level")["hard_hit_pct_allowed"].transform(zscore)
        df["_z_barrel_allowed_inv"] = -df.groupby("level")["barrel_pct_approx_allowed"].transform(zscore)
        skill_cols += ["_z_ev_allowed_inv", "_z_hard_hit_allowed_inv", "_z_barrel_allowed_inv"]

    df["_z_era_inv"] = -df.groupby("level")["era"].transform(lambda s: zscore(pd.to_numeric(s, errors="coerce")))
    results_cols = ["_z_era_inv"]

    df["skill_z"] = df.apply(lambda r: _row_mean_ignore_none(r, skill_cols), axis=1)
    df["results_z"] = df.apply(lambda r: _row_mean_ignore_none(r, results_cols), axis=1)
    df["undervalued_score"] = df.apply(
        lambda r: (r["skill_z"] - r["results_z"])
        if r["skill_z"] is not None and r["results_z"] is not None else None,
        axis=1)
    df["has_statcast"] = "avg_exit_velocity_allowed" in df.columns and df["avg_exit_velocity_allowed"].notna()

    df = df.drop(columns=[c for c in df.columns if c.startswith("_z_")])
    return df


def approx_barrel(row):
    ev = row["launch_speed"]
    la = row["launch_angle"]
    if pd.isna(ev) or pd.isna(la):
        return False
    if ev >= 98 and 26 <= la <= 30:
        return True
    if ev > 98:
        extra = min((ev - 98) * 0.5, 10)
        return (26 - extra) <= la <= (30 + extra)
    return False


def summarize_statcast(raw_df, id_col="playerId", name_col="playerName", suffix=""):
    """id_col/name_col let this summarize either the batter's own contact
    (playerId/playerName) or the pitcher's contact ALLOWED (pitcherId/
    pitcherName) from the same raw batted-ball rows. suffix distinguishes
    the resulting column names (e.g. '_allowed' for pitchers)."""
    if raw_df is None or raw_df.empty or "launch_speed" not in raw_df.columns:
        return pd.DataFrame()
    bb = raw_df.dropna(subset=["launch_speed", id_col])
    if bb.empty:
        return pd.DataFrame()
    bb = bb.copy()
    bb["is_barrel"] = bb.apply(approx_barrel, axis=1)

    def player_agg(g):
        n = len(g)
        hard_hit = (g["launch_speed"] >= 95).sum()
        return pd.Series({
            f"statcast_BBE{suffix}": n,
            f"avg_exit_velocity{suffix}": g["launch_speed"].mean(),
            f"max_exit_velocity{suffix}": g["launch_speed"].max(),
            f"avg_launch_angle{suffix}": g["launch_angle"].mean(),
            f"hard_hit_pct{suffix}": hard_hit / n if n else None,
            f"barrel_pct_approx{suffix}": g["is_barrel"].sum() / n if n else None,
        })

    summary = bb.groupby(id_col).apply(player_agg).reset_index()
    summary = summary.rename(columns={id_col: "playerId"})
    return summary


def merge_statcast(df, statcast_summary):
    if df.empty or statcast_summary is None or statcast_summary.empty:
        return df
    return df.merge(statcast_summary, on="playerId", how="left", suffixes=("", "_sc"))


# ---------------------------------------------------------------------------
# Pitch arsenal, control, and stuff grades
#
# Pitch-level tracking (velocity, spin rate, movement) via TrackMan is
# rolled out more broadly across full-season affiliates than full
# Hawk-Eye batted-ball Statcast, but coverage still isn't guaranteed at
# every level (and likely doesn't exist at complex-level Rookie/DSL ball).
# Rather than assume which levels have it, this pulls pitch-by-pitch data
# for every team and lets whatever actually comes back determine
# availability -- spin rate in particular may be missing even where
# velocity is present.
# ---------------------------------------------------------------------------

FASTBALL_CODES = {"FF", "FT", "SI", "FC"}
WHIFF_DESCRIPTIONS = {"Swinging Strike", "Swinging Strike (Blocked)"}
SWING_DESCRIPTIONS = WHIFF_DESCRIPTIONS | {
    "Foul", "Foul Tip", "Foul Bunt", "Missed Bunt",
    "In play, out(s)", "In play, run(s)", "In play, no out",
}
IN_ZONE_CODES = {1, 2, 3, 4, 5, 6, 7, 8, 9}


def summarize_arsenal(pitches_df):
    """One row per pitcher per pitch type: usage%, velocity, spin (if
    tracked), movement, and whiff rate. This is the actual 'arsenal'
    breakdown -- the per-pitcher, per-pitch-type table."""
    if pitches_df is None or pitches_df.empty or "pitch_type" not in pitches_df.columns:
        return pd.DataFrame()

    df = pitches_df.dropna(subset=["pitcherId", "pitch_type"]).copy()
    if df.empty:
        return pd.DataFrame()

    df["is_swing"] = df["call_description"].isin(SWING_DESCRIPTIONS)
    df["is_whiff"] = df["call_description"].isin(WHIFF_DESCRIPTIONS)

    totals = df.groupby("pitcherId").size().rename("total_pitches")

    def agg(g):
        n = len(g)
        swings = g["is_swing"].sum()
        whiffs = g["is_whiff"].sum()
        return pd.Series({
            "pitcherName": g["pitcherName"].iloc[0],
            "pitches_thrown": n,
            "avg_velo": g["start_speed"].mean(),
            "max_velo": g["start_speed"].max(),
            "avg_spin_rate": g["spin_rate"].mean() if g["spin_rate"].notna().any() else None,
            "avg_horiz_break_in": (g["pfx_x"].mean() * 12) if g["pfx_x"].notna().any() else None,
            "avg_vert_break_in": (g["pfx_z"].mean() * 12) if g["pfx_z"].notna().any() else None,
            "whiff_pct": (whiffs / swings) if swings >= 5 else None,
        })

    arsenal = df.groupby(["pitcherId", "pitch_type"]).apply(agg).reset_index()
    arsenal = arsenal.merge(totals, on="pitcherId", how="left")
    arsenal["usage_pct"] = arsenal["pitches_thrown"] / arsenal["total_pitches"]
    arsenal = arsenal.drop(columns=["total_pitches"])
    return arsenal.sort_values(["pitcherId", "usage_pct"], ascending=[True, False])


def summarize_pitcher_rollup(pitches_df):
    """One row per pitcher: zone%, overall whiff%, and primary fastball
    velo/spin -- the rollup numbers used to build Control/Stuff grades."""
    if pitches_df is None or pitches_df.empty:
        return pd.DataFrame()

    df = pitches_df.dropna(subset=["pitcherId"]).copy()
    if df.empty:
        return pd.DataFrame()

    df["is_swing"] = df["call_description"].isin(SWING_DESCRIPTIONS)
    df["is_whiff"] = df["call_description"].isin(WHIFF_DESCRIPTIONS)
    df["in_zone"] = df["zone"].apply(lambda z: z in IN_ZONE_CODES if pd.notna(z) else None)

    rows = []
    for pid, g in df.groupby("pitcherId"):
        swings = g["is_swing"].sum()
        whiffs = g["is_whiff"].sum()
        zone_known = g["in_zone"].notna().sum()

        fb = g[g["pitch_type"].isin(FASTBALL_CODES)]
        primary_fb_velo, primary_fb_spin, primary_fb_type = None, None, None
        if not fb.empty:
            usage_by_type = fb.groupby("pitch_type").size()
            primary_fb_type = usage_by_type.idxmax()
            fb_primary = fb[fb["pitch_type"] == primary_fb_type]
            primary_fb_velo = fb_primary["start_speed"].mean()
            primary_fb_spin = fb_primary["spin_rate"].mean() if fb_primary["spin_rate"].notna().any() else None

        rows.append({
            "playerId": pid,
            "zone_pct": (g["in_zone"].sum() / zone_known) if zone_known else None,
            "overall_whiff_pct": (whiffs / swings) if swings >= 10 else None,
            "primary_fastball_type": primary_fb_type,
            "primary_fastball_velo": primary_fb_velo,
            "primary_fastball_spin": primary_fb_spin,
            "has_pitch_tracking": g["start_speed"].notna().any(),
            "has_spin_tracking": g["spin_rate"].notna().any(),
        })
    return pd.DataFrame(rows)


def summarize_hitter_approach(pitches_df):
    """Real swing-decision data from the batter's side of the same
    pitch-level feed: chase% (swinging at pitches out of the zone),
    swinging-strike% (whiffs per pitch seen), and zone-contact% (contact
    rate when swinging at pitches in the zone). Zone/call data is base
    Gameday infrastructure, not Statcast-exclusive, so this should have
    real coverage well beyond just the Statcast-tracked levels -- but as
    with everything else here, coverage is determined by what actually
    comes back, not assumed in advance."""
    if pitches_df is None or pitches_df.empty or "batterId" not in pitches_df.columns:
        return pd.DataFrame()

    df = pitches_df.dropna(subset=["batterId"]).copy()
    if df.empty:
        return pd.DataFrame()

    df["is_swing"] = df["call_description"].isin(SWING_DESCRIPTIONS)
    df["is_whiff"] = df["call_description"].isin(WHIFF_DESCRIPTIONS)
    df["in_zone"] = df["zone"].apply(lambda z: z in IN_ZONE_CODES if pd.notna(z) else None)

    rows = []
    for bid, g in df.groupby("batterId"):
        n = len(g)
        swings = g["is_swing"].sum()
        whiffs = g["is_whiff"].sum()
        zone_known = g["in_zone"].notna().sum()
        out_of_zone = (g["in_zone"] == False).sum()  # noqa: E712
        chase_swings = ((g["is_swing"]) & (g["in_zone"] == False)).sum()  # noqa: E712
        zone_swings = ((g["is_swing"]) & (g["in_zone"] == True)).sum()  # noqa: E712
        zone_whiffs = ((g["is_whiff"]) & (g["in_zone"] == True)).sum()  # noqa: E712

        rows.append({
            "playerId": bid,
            "pitches_seen": n,
            "swing_pct": (swings / n) if n else None,
            "batter_whiff_pct": (whiffs / swings) if swings >= 10 else None,
            "chase_pct": (chase_swings / out_of_zone) if out_of_zone >= 10 else None,
            "zone_contact_pct": ((zone_swings - zone_whiffs) / zone_swings) if zone_swings >= 10 else None,
            "has_pitch_tracking_batter": zone_known > 0,
        })
    return pd.DataFrame(rows)


def compute_pitcher_grades(pitching):
    """Adds control_grade, stuff_grade, and pitching_overall_fv (20-80)
    to the pitching dataframe, using whatever pitch-tracking data is
    actually present. Confidence notes are added since spin tracking in
    particular may not exist everywhere."""
    if pitching.empty:
        return pitching

    df = pitching.copy()

    # --- Control: zone% + BB% (inverted)
    control_cols = []
    if "zone_pct" in df.columns:
        df["_z_zone"] = df.groupby("level")["zone_pct"].transform(zscore)
        control_cols.append("_z_zone")
    df["_z_bb_inv"] = -df.groupby("level")["BB_pct"].transform(zscore)
    control_cols.append("_z_bb_inv")
    df["_control_z"] = df.apply(lambda r: _row_mean_ignore_none(r, control_cols), axis=1)
    df["control_grade"] = df["_control_z"].apply(to_20_80)
    df["control_grade_confidence"] = (
        "medium (zone% + BB%)" if "zone_pct" in df.columns
        else "low (BB% only -- zone data unavailable)"
    )

    # --- Stuff: primary fastball velo/spin (if tracked) + overall whiff%
    stuff_cols = []
    if "primary_fastball_velo" in df.columns:
        df["_z_fbvelo"] = df.groupby(["level", "primary_fastball_type"])["primary_fastball_velo"].transform(zscore)
        stuff_cols.append("_z_fbvelo")
        if df.get("has_spin_tracking", pd.Series(dtype=bool)).any():
            df["_z_fbspin"] = df.groupby(["level", "primary_fastball_type"])["primary_fastball_spin"].transform(zscore)
            stuff_cols.append("_z_fbspin")
    if "overall_whiff_pct" in df.columns:
        df["_z_whiff"] = df.groupby("level")["overall_whiff_pct"].transform(zscore)
        stuff_cols.append("_z_whiff")

    if stuff_cols:
        df["_stuff_z"] = df.apply(lambda r: _row_mean_ignore_none(r, stuff_cols), axis=1)
        df["stuff_grade"] = df["_stuff_z"].apply(to_20_80)
        df["stuff_grade_confidence"] = (
            "medium-high (velo + spin + whiff%)" if "_z_fbspin" in df.columns
            else "medium (velo + whiff%, no spin tracking at this level)"
        )
    else:
        df["stuff_grade"] = None
        df["stuff_grade_confidence"] = "unavailable (no pitch tracking at this level)"

    def overall_fv(row):
        available = {
            "control": row.get("control_grade"),
            "stuff": row.get("stuff_grade"),
        }
        available = {k: v for k, v in available.items() if pd.notna(v)}
        if not available:
            return None
        weights = {"control": 0.5, "stuff": 0.5}
        total_weight = sum(weights[k] for k in available)
        return round(sum(weights[k] * v for k, v in available.items()) / total_weight)

    df["pitching_overall_fv"] = df.apply(overall_fv, axis=1)

    df = df.drop(columns=[c for c in df.columns if c.startswith("_")])
    return df


# ---------------------------------------------------------------------------
# Full pull (box score, all levels) + Statcast (tracked levels only),
# fully combined in one call
# ---------------------------------------------------------------------------

def build_full(season, progress_callback=None):
    """progress_callback(pct: float 0-1, msg: str) -> None, optional."""

    def report(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
        else:
            print(f"[{pct:5.1%}] {msg}")

    report(0.0, "Fetching Mets affiliate teams...")
    teams = get_affiliate_teams(season)

    hitting_frames, pitching_frames, fielding_frames = [], [], []
    n_teams = max(len(teams), 1)
    for i, t in enumerate(teams):
        report((i / n_teams) * 0.35, f"Box score stats: {t['teamName']} ({t['levelName']})...")
        hrows = get_team_group_stats(t["teamId"], t["sportId"], season, "hitting")
        for r in hrows:
            r["level"] = t["levelName"]
            r["team"] = t["teamName"]
        hitting_frames.append(pd.DataFrame(hrows))

        prows = get_team_group_stats(t["teamId"], t["sportId"], season, "pitching")
        for r in prows:
            r["level"] = t["levelName"]
            r["team"] = t["teamName"]
        pitching_frames.append(pd.DataFrame(prows))

        frows = get_team_fielding_stats(t["teamId"], t["sportId"], season)
        fielding_frames.extend(frows)
        time.sleep(0.15)

    hitting = pd.concat(hitting_frames, ignore_index=True) if hitting_frames else pd.DataFrame()
    pitching = pd.concat(pitching_frames, ignore_index=True) if pitching_frames else pd.DataFrame()
    fielding_agg = aggregate_fielding(fielding_frames)

    # Bio / age enrichment
    report(0.35, "Fetching player bios/ages...")
    all_ids = set()
    if not hitting.empty:
        all_ids |= set(hitting["playerId"].dropna().unique())
    if not pitching.empty:
        all_ids |= set(pitching["playerId"].dropna().unique())
    all_ids = list(all_ids)

    bios = {}
    for i, pid in enumerate(all_ids):
        if i % 15 == 0:
            report(0.35 + (i / max(len(all_ids), 1)) * 0.25, f"Bios ({i}/{len(all_ids)})...")
        bios[pid] = get_player_bio(pid)
        time.sleep(0.04)

    bio_df = pd.DataFrame.from_dict(bios, orient="index")
    bio_df.index.name = "playerId"
    bio_df.reset_index(inplace=True)

    if not hitting.empty:
        hitting = hitting.merge(bio_df, on="playerId", how="left")
        hitting = add_hitting_derived(hitting)
        if not fielding_agg.empty:
            hitting = hitting.merge(fielding_agg, on="playerId", how="left")

    if not pitching.empty:
        pitching = pitching.merge(bio_df, on="playerId", how="left")
        pitching = add_pitching_derived(pitching)

    # Per-game data (batted balls + pitch-level tracking) -- pulled for
    # EVERY team/level now, not just the ones known to have full Statcast.
    # Whether a given level actually has tracking shows up in the data
    # itself (NaN columns) rather than being assumed in advance.
    batter_summaries, pitcher_bb_summaries, all_pitches = [], [], []
    n_teams2 = max(len(teams), 1)
    for i, t in enumerate(teams):
        base_pct = 0.6 + (i / n_teams2) * 0.38

        def cb(msg, base_pct=base_pct):
            report(base_pct, f"Game data: {msg}")

        bb, pitches = pull_game_data_for_team(t["teamId"], t["sportId"], season,
                                               progress_callback=cb, label=t["teamName"])
        b_summary = summarize_statcast(bb, id_col="playerId", name_col="playerName")
        if not b_summary.empty:
            batter_summaries.append(b_summary)
        p_summary = summarize_statcast(bb, id_col="pitcherId", name_col="pitcherName", suffix="_allowed")
        if not p_summary.empty:
            pitcher_bb_summaries.append(p_summary)
        if not pitches.empty:
            all_pitches.append(pitches)

    if batter_summaries and not hitting.empty:
        combined = pd.concat(batter_summaries, ignore_index=True).drop_duplicates(subset=["playerId"])
        hitting = merge_statcast(hitting, combined)
    if pitcher_bb_summaries and not pitching.empty:
        combined = pd.concat(pitcher_bb_summaries, ignore_index=True).drop_duplicates(subset=["playerId"])
        pitching = merge_statcast(pitching, combined)

    all_pitches_df = pd.concat(all_pitches, ignore_index=True) if all_pitches else pd.DataFrame()
    arsenal = summarize_arsenal(all_pitches_df)
    rollup = summarize_pitcher_rollup(all_pitches_df)
    if not rollup.empty and not pitching.empty:
        pitching = pitching.merge(rollup, on="playerId", how="left")

    hitter_approach = summarize_hitter_approach(all_pitches_df)
    if not hitter_approach.empty and not hitting.empty:
        hitting = hitting.merge(hitter_approach, on="playerId", how="left")

    # Now compute scores/grades -- after everything above is merged in,
    # so these functions can detect and use whatever's actually present.
    if not hitting.empty:
        hitting = compute_hitting_scores(hitting)
        hitting = compute_tool_grades(hitting)
    if not pitching.empty:
        pitching = compute_pitching_scores(pitching)
        pitching = compute_pitcher_grades(pitching)

    report(1.0, "Done.")
    return hitting, pitching, arsenal, teams


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

INVALID_SHEET_CHARS = ['\\', '/', '?', '*', '[', ']', ':']


def safe_sheet_name(name, used_names):
    cleaned = name
    for ch in INVALID_SHEET_CHARS:
        cleaned = cleaned.replace(ch, "-")
    cleaned = cleaned.strip().strip("'")
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 1
    while cleaned.lower() in used_names:
        suffix_str = f"_{suffix}"
        cleaned = (base[: 31 - len(suffix_str)] + suffix_str)
        suffix += 1
    used_names.add(cleaned.lower())
    return cleaned


def write_excel_bytes(hitting, pitching, teams, arsenal=None):
    level_order = list(dict.fromkeys(t["levelName"] for t in teams))
    buffer = io.BytesIO()
    used_names = set()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if not hitting.empty:
            name = safe_sheet_name("Hitting - All Levels", used_names)
            hitting.sort_values("undervalued_score", ascending=False).to_excel(
                writer, sheet_name=name, index=False)
        if not pitching.empty:
            name = safe_sheet_name("Pitching - All Levels", used_names)
            pitching.sort_values("undervalued_score", ascending=False).to_excel(
                writer, sheet_name=name, index=False)
        if arsenal is not None and not arsenal.empty:
            name = safe_sheet_name("Pitch Arsenal", used_names)
            arsenal.sort_values(["pitcherId", "usage_pct"], ascending=[True, False]).to_excel(
                writer, sheet_name=name, index=False)

        for level in level_order:
            if not hitting.empty:
                sub = hitting[hitting["level"] == level].sort_values("undervalued_score", ascending=False)
                if not sub.empty:
                    name = safe_sheet_name(f"Hit-{level}", used_names)
                    sub.to_excel(writer, sheet_name=name, index=False)
            if not pitching.empty:
                sub = pitching[pitching["level"] == level].sort_values("undervalued_score", ascending=False)
                if not sub.empty:
                    name = safe_sheet_name(f"Pitch-{level}", used_names)
                    sub.to_excel(writer, sheet_name=name, index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 1:
            continue
        last_col = get_column_letter(ws.max_column)
        ref = f"A1:{last_col}{ws.max_row}"
        table_name = "T_" + "".join(ch for ch in ws.title if ch.isalnum())
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer
