"""
streamlit_app.py

Streamlit app: pulls EVERY Mets minor-league affiliate's hitting and
pitching season stats (no major leaguers) from the public MLB Stats API,
computes undervaluation metrics, and lets you download a sortable,
level-broken-out Excel workbook.

requirements.txt should contain:
    streamlit
    requests
    pandas
    openpyxl
"""

import io
import time

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = "https://statsapi.mlb.com/api/v1"
METS_ORG_ID = 121

SPORT_IDS = {
    11: "Triple-A",
    12: "Double-A",
    13: "High-A",
    14: "Single-A",
    16: "Rookie/Complex (FCL)",
    5442: "Dominican Summer League",
}

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "personal-research-app/1.0"})


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def get_json(url, params=None, retries=3, sleep=0.4):
    for attempt in range(retries):
        try:
            r = SESSION.get(url, params=params, timeout=20)
            r.raise_for_status()
            return r.json()
        except Exception:
            if attempt == retries - 1:
                return {}
            time.sleep(sleep)
    return {}


def get_affiliate_teams(season):
    teams = []
    for sport_id, level_name in SPORT_IDS.items():
        data = get_json(f"{BASE}/teams", params={"sportId": sport_id, "season": season})
        for t in data.get("teams", []):
            if t.get("parentOrgId") == METS_ORG_ID:
                teams.append({
                    "teamId": t["id"],
                    "teamName": t.get("name"),
                    "sportId": sport_id,
                    "levelName": level_name,
                })
    return teams


def get_team_group_stats(team_id, sport_id, season, group):
    data = get_json(f"{BASE}/stats", params={
        "stats": "season",
        "group": group,
        "season": season,
        "sportId": sport_id,
        "teamId": team_id,
        "limit": 300,
    })
    rows = []
    for block in data.get("stats", []):
        for split in block.get("splits", []):
            player = split.get("player", {})
            stat = split.get("stat", {})
            row = {"playerId": player.get("id"), "playerName": player.get("fullName")}
            row.update(stat)
            rows.append(row)
    return rows


def get_player_bio(player_id):
    data = get_json(f"{BASE}/people/{player_id}")
    people = data.get("people", [])
    if not people:
        return {}
    p = people[0]
    return {
        "birthDate": p.get("birthDate"),
        "currentAge": p.get("currentAge"),
        "position": (p.get("primaryPosition") or {}).get("abbreviation"),
        "bats": (p.get("batSide") or {}).get("code"),
        "throws": (p.get("pitchHand") or {}).get("code"),
        "height": p.get("height"),
        "weight": p.get("weight"),
    }


# ---------------------------------------------------------------------------
# Derived stats
# ---------------------------------------------------------------------------

def safe_div(n, d):
    try:
        n = float(n)
        d = float(d)
        return n / d if d else None
    except (TypeError, ValueError):
        return None


def add_hitting_derived(df):
    if df.empty:
        return df
    for col in ["hits", "homeRuns", "atBats", "strikeOuts", "baseOnBalls",
                "sacFlies", "plateAppearances"]:
        if col not in df.columns:
            df[col] = 0
    df["BABIP"] = df.apply(
        lambda r: safe_div(
            (float(r["hits"] or 0) - float(r["homeRuns"] or 0)),
            (float(r["atBats"] or 0) - float(r["strikeOuts"] or 0)
             - float(r["homeRuns"] or 0) + float(r["sacFlies"] or 0)),
        ), axis=1)
    df["K_pct"] = df.apply(lambda r: safe_div(r["strikeOuts"], r["plateAppearances"]), axis=1)
    df["BB_pct"] = df.apply(lambda r: safe_div(r["baseOnBalls"], r["plateAppearances"]), axis=1)
    if "slg" in df.columns and "avg" in df.columns:
        df["ISO"] = df.apply(
            lambda r: (float(r["slg"]) - float(r["avg"]))
            if r.get("slg") not in (None, "", "-") and r.get("avg") not in (None, "", "-")
            else None, axis=1)
    return df


def add_pitching_derived(df):
    if df.empty:
        return df
    for col in ["strikeOuts", "baseOnBalls", "battersFaced"]:
        if col not in df.columns:
            df[col] = 0
    df["K_pct"] = df.apply(lambda r: safe_div(r["strikeOuts"], r["battersFaced"]), axis=1)
    df["BB_pct"] = df.apply(lambda r: safe_div(r["baseOnBalls"], r["battersFaced"]), axis=1)
    df["K_minus_BB_pct"] = df.apply(
        lambda r: (r["K_pct"] - r["BB_pct"]) if r["K_pct"] is not None and r["BB_pct"] is not None else None,
        axis=1)
    return df


# ---------------------------------------------------------------------------
# Statcast (minors) via MLB Stats API play-by-play feed
#
# Public Statcast tracking only exists for Triple-A (since 2023) and the
# Florida State League / Single-A (since 2021). For games at those levels,
# the MLB Stats API's own play-by-play feed includes a "hitData" block on
# batted-ball events with launchSpeed / launchAngle / totalDistance -- this
# is the same underlying tracking data, exposed through the documented,
# reliable Stats API instead of an undocumented Savant CSV endpoint.
#
# This pulls every regular-season game for the given team, fetches play-by-
# play for each, and aggregates batted-ball metrics per player. It's more
# API calls than the season-stat endpoints (one per game), so it's slower,
# but it doesn't depend on guessing at an unstable scraping target.
# ---------------------------------------------------------------------------

STATCAST_TRACKED_LEVELS = {"Triple-A", "Single-A"}


def get_team_game_pks(team_id, sport_id, season):
    data = get_json(f"{BASE}/schedule", params={
        "teamId": team_id,
        "sportId": sport_id,
        "season": season,
        "gameType": "R",
    })
    pks = []
    for date_block in data.get("dates", []):
        for game in date_block.get("games", []):
            if game.get("status", {}).get("abstractGameState") == "Final":
                pks.append(game["gamePk"])
    return pks


def get_batted_balls_for_game(game_pk):
    """Returns list of dicts: playerId, playerName, launchSpeed, launchAngle,
    totalDistance for every tracked batted ball in this game."""
    data = get_json(f"{BASE}/game/{game_pk}/playByPlay")
    rows = []
    for play in data.get("allPlays", []):
        matchup = play.get("matchup", {})
        batter = matchup.get("batter", {})
        for event in play.get("playEvents", []):
            hit_data = event.get("hitData")
            if hit_data and hit_data.get("launchSpeed") is not None:
                rows.append({
                    "playerId": batter.get("id"),
                    "playerName": batter.get("fullName"),
                    "launch_speed": hit_data.get("launchSpeed"),
                    "launch_angle": hit_data.get("launchAngle"),
                    "total_distance": hit_data.get("totalDistance"),
                })
    return rows


def pull_statcast_for_team(team_id, sport_id, season, progress_callback=None, label=""):
    game_pks = get_team_game_pks(team_id, sport_id, season)
    all_rows = []
    for i, pk in enumerate(game_pks):
        if progress_callback and i % 5 == 0:
            progress_callback(f"{label}: game {i+1}/{len(game_pks)}...")
        all_rows.extend(get_batted_balls_for_game(pk))
        time.sleep(0.1)
    return pd.DataFrame(all_rows)


def summarize_statcast(raw_df):
    """Collapse batted-ball-level rows into one row per player with
    headline metrics: EV, hard-hit%, approx barrel%, etc."""
    if raw_df is None or raw_df.empty or "launch_speed" not in raw_df.columns:
        return pd.DataFrame()

    bb = raw_df.dropna(subset=["launch_speed"])
    if bb.empty:
        return pd.DataFrame()

    def approx_barrel(row):
        # Simplified barrel approximation (real Savant barrel classification
        # uses a speed/angle matrix; this is a reasonable stand-in since the
        # exact matrix isn't publicly documented in closed form).
        ev = row["launch_speed"]
        la = row["launch_angle"]
        if pd.isna(ev) or pd.isna(la):
            return False
        if ev >= 98 and 26 <= la <= 30:
            return True
        if ev > 98:
            # widen angle window slightly as EV climbs, mirroring Savant's
            # real barrel definition shape
            extra = min((ev - 98) * 0.5, 10)
            return (26 - extra) <= la <= (30 + extra)
        return False

    bb = bb.copy()
    bb["is_barrel"] = bb.apply(approx_barrel, axis=1)

    def player_agg(g):
        n = len(g)
        hard_hit = (g["launch_speed"] >= 95).sum()
        return pd.Series({
            "statcast_BBE": n,
            "avg_exit_velocity": g["launch_speed"].mean(),
            "max_exit_velocity": g["launch_speed"].max(),
            "avg_launch_angle": g["launch_angle"].mean(),
            "hard_hit_pct": hard_hit / n if n else None,
            "barrel_pct_approx": g["is_barrel"].sum() / n if n else None,
        })

    summary = bb.groupby("playerId").apply(player_agg).reset_index()
    return summary


def merge_statcast(df, statcast_summary):
    if df.empty or statcast_summary is None or statcast_summary.empty:
        return df
    return df.merge(statcast_summary, on="playerId", how="left", suffixes=("", "_sc"))


def zscore(series):
    s = pd.to_numeric(series, errors="coerce")
    if s.dropna().empty or s.std(skipna=True) in (0, None):
        return pd.Series([None] * len(series), index=series.index)
    return (s - s.mean(skipna=True)) / s.std(skipna=True)


# ---------------------------------------------------------------------------
# Pull + build
# ---------------------------------------------------------------------------

def build(season, progress_callback=None):
    teams = get_affiliate_teams(season)
    hitting_frames, pitching_frames = [], []

    total_steps = max(len(teams), 1)
    for i, t in enumerate(teams):
        if progress_callback:
            progress_callback((i / total_steps) * 0.6, f"Pulling {t['teamName']} ({t['levelName']})...")

        hrows = get_team_group_stats(t["teamId"], t["sportId"], season, "hitting")
        for r in hrows:
            r["level"] = t["levelName"]
            r["team"] = t["teamName"]
        hitting_frames.append(pd.DataFrame(hrows))

        prows = get_team_group_stats(t["teamId"], t["sportId"], season, "pitching")
        for r in prows:
            r["level"] = t["levelName"]
            r["team"] = t["teamName"]
        pitching_frames.append(pd.DataFrame(prows))
        time.sleep(0.15)

    hitting = pd.concat(hitting_frames, ignore_index=True) if hitting_frames else pd.DataFrame()
    pitching = pd.concat(pitching_frames, ignore_index=True) if pitching_frames else pd.DataFrame()

    all_ids = set()
    if not hitting.empty:
        all_ids |= set(hitting["playerId"].dropna().unique())
    if not pitching.empty:
        all_ids |= set(pitching["playerId"].dropna().unique())

    bios = {}
    all_ids = list(all_ids)
    for i, pid in enumerate(all_ids):
        if progress_callback and i % 10 == 0:
            progress_callback(0.6 + (i / max(len(all_ids), 1)) * 0.35, f"Fetching player bios ({i}/{len(all_ids)})...")
        bios[pid] = get_player_bio(pid)
        time.sleep(0.04)

    bio_df = pd.DataFrame.from_dict(bios, orient="index")
    bio_df.index.name = "playerId"
    bio_df.reset_index(inplace=True)

    if not hitting.empty:
        hitting = hitting.merge(bio_df, on="playerId", how="left")
        hitting = add_hitting_derived(hitting)
        hitting["age_vs_level_z"] = hitting.groupby("level")["currentAge"].transform(zscore)
        hitting["BABIP_z_in_level"] = hitting.groupby("level")["BABIP"].transform(zscore)
        hitting["K_pct_z_in_level"] = hitting.groupby("level")["K_pct"].transform(zscore)
        hitting["BB_pct_z_in_level"] = hitting.groupby("level")["BB_pct"].transform(zscore)
        hitting["undervalued_score"] = (
            -hitting["age_vs_level_z"].fillna(0)
            + hitting["BB_pct_z_in_level"].fillna(0)
            - hitting["K_pct_z_in_level"].fillna(0)
            - hitting["BABIP_z_in_level"].fillna(0)
        )

    if not pitching.empty:
        pitching = pitching.merge(bio_df, on="playerId", how="left")
        pitching = add_pitching_derived(pitching)
        pitching["age_vs_level_z"] = pitching.groupby("level")["currentAge"].transform(zscore)
        pitching["K_minus_BB_z_in_level"] = pitching.groupby("level")["K_minus_BB_pct"].transform(zscore)
        pitching["ERA_z_in_level"] = pitching.groupby("level")["era"].transform(
            lambda s: zscore(pd.to_numeric(s, errors="coerce")))
        pitching["undervalued_score"] = (
            -pitching["age_vs_level_z"].fillna(0)
            + pitching["K_minus_BB_z_in_level"].fillna(0)
            + pitching["ERA_z_in_level"].fillna(0)
        )

    if progress_callback:
        progress_callback(1.0, "Done.")

    return hitting, pitching, teams


INVALID_SHEET_CHARS = ['\\', '/', '?', '*', '[', ']', ':']


def safe_sheet_name(name, used_names):
    cleaned = name
    for ch in INVALID_SHEET_CHARS:
        cleaned = cleaned.replace(ch, "-")
    cleaned = cleaned.strip().strip("'")
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 1
    while cleaned.lower() in used_names:
        suffix_str = f"_{suffix}"
        cleaned = (base[: 31 - len(suffix_str)] + suffix_str)
        suffix += 1
    used_names.add(cleaned.lower())
    return cleaned


def write_excel_bytes(hitting, pitching, teams):
    level_order = list(dict.fromkeys(t["levelName"] for t in teams))
    buffer = io.BytesIO()
    used_names = set()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if not hitting.empty:
            name = safe_sheet_name("Hitting - All Levels", used_names)
            hitting.sort_values("undervalued_score", ascending=False).to_excel(
                writer, sheet_name=name, index=False)
        if not pitching.empty:
            name = safe_sheet_name("Pitching - All Levels", used_names)
            pitching.sort_values("undervalued_score", ascending=False).to_excel(
                writer, sheet_name=name, index=False)

        for level in level_order:
            if not hitting.empty:
                sub = hitting[hitting["level"] == level].sort_values("undervalued_score", ascending=False)
                if not sub.empty:
                    name = safe_sheet_name(f"Hit-{level}", used_names)
                    sub.to_excel(writer, sheet_name=name, index=False)
            if not pitching.empty:
                sub = pitching[pitching["level"] == level].sort_values("undervalued_score", ascending=False)
                if not sub.empty:
                    name = safe_sheet_name(f"Pitch-{level}", used_names)
                    sub.to_excel(writer, sheet_name=name, index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 1:
            continue
        last_col = get_column_letter(ws.max_column)
        ref = f"A1:{last_col}{ws.max_row}"
        table_name = "T_" + "".join(ch for ch in ws.title if ch.isalnum())
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Mets MiLB Stat Pull", layout="wide")
st.title("⚾ Mets Minor League Stat Pull")
st.caption("Pulls every Mets affiliate (Syracuse, Binghamton, Brooklyn, St. Lucie, FCL, DSL) — no major leaguers — and flags players whose underlying stats outpace their results.")

season = st.number_input("Season", min_value=2015, max_value=2030, value=2026, step=1)

if "result" not in st.session_state:
    st.session_state.result = None

if st.button("Pull data", type="primary"):
    progress_bar = st.progress(0.0)
    status = st.empty()

    def progress_callback(pct, msg):
        progress_bar.progress(min(max(pct, 0.0), 1.0))
        status.text(msg)

    with st.spinner("Working..."):
        hitting, pitching, teams = build(int(season), progress_callback)

    st.session_state.result = {
        "hitting": hitting,
        "pitching": pitching,
        "teams": teams,
    }
    status.text("Done.")

if st.session_state.result:
    res = st.session_state.result
    st.success(f"Pulled {len(res['hitting'])} hitter-rows and {len(res['pitching'])} pitcher-rows across {len(res['teams'])} teams.")

    st.divider()
    st.subheader("Statcast (Triple-A & Single-A only)")
    st.caption(
        "Public Statcast tracking for the minors only exists for Triple-A "
        "(Syracuse) since 2023 and the Florida State League / Single-A "
        "(St. Lucie) since 2021. Double-A, High-A, Rookie, and DSL have no "
        "public Statcast anywhere -- that's a real data gap. This pulls "
        "exit velocity / launch angle / hard-hit% directly from the MLB "
        "Stats API's game-by-game play-by-play feed (the same documented "
        "API used above), aggregated per player. It's slow because it "
        "fetches every game individually -- expect a few minutes."
    )

    if st.button("Pull Statcast (Syracuse + St. Lucie)"):
        tracked_teams = [t for t in res["teams"] if t["levelName"] in STATCAST_TRACKED_LEVELS]
        if not tracked_teams:
            st.warning("No Triple-A or Single-A team found in the pulled data.")
        else:
            sc_progress = st.progress(0.0)
            sc_status = st.empty()
            all_summaries = []
            for i, t in enumerate(tracked_teams):
                def cb(msg):
                    sc_status.text(msg)
                with st.spinner(f"Pulling Statcast for {t['teamName']}..."):
                    raw = pull_statcast_for_team(
                        t["teamId"], t["sportId"], int(season),
                        progress_callback=cb, label=t["teamName"])
                    summary = summarize_statcast(raw)
                    summary["level"] = t["levelName"]
                    all_summaries.append(summary)
                sc_progress.progress((i + 1) / len(tracked_teams))

            combined_summary = pd.concat(all_summaries, ignore_index=True) if all_summaries else pd.DataFrame()
            if combined_summary.empty:
                st.warning(
                    "No batted-ball tracking data came back for these games. "
                    "This can happen if a level's tracking coverage changed "
                    "or games aren't fully loaded yet -- try again later in "
                    "the season, or double check the season year."
                )
            else:
                res["hitting"] = merge_statcast(res["hitting"], combined_summary.drop(columns=["level"]))
                st.session_state.result = res
                sc_status.text("Done.")
                st.success(f"Merged Statcast data for {len(combined_summary)} players.")

    st.divider()
    excel_bytes = write_excel_bytes(res["hitting"], res["pitching"], res["teams"])
    st.download_button(
        label="📥 Download Excel workbook",
        data=excel_bytes,
        file_name=f"mets_milb_{season}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    tab1, tab2 = st.tabs(["Hitting preview", "Pitching preview"])
    with tab1:
        if not res["hitting"].empty:
            st.dataframe(
                res["hitting"].sort_values("undervalued_score", ascending=False),
                use_container_width=True,
            )
        else:
            st.info("No hitting data returned.")
    with tab2:
        if not res["pitching"].empty:
            st.dataframe(
                res["pitching"].sort_values("undervalued_score", ascending=False),
                use_container_width=True,
            )
        else:
            st.info("No pitching data returned.")
